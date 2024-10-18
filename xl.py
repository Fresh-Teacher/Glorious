import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, PieChart

def create_chart(sheet, chart_type, title, start_cell):
    if chart_type == "bar":
        chart = BarChart()
        chart.y_axis.title = "Rainfall (mm)"
        chart.x_axis.title = "Month"
    elif chart_type == "pie":
        chart = PieChart()
    
    data = Reference(sheet, min_col=2, min_row=3, max_row=9)
    categories = Reference(sheet, min_col=1, min_row=4, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    sheet.add_chart(chart, start_cell)

def create_rainfall_excel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Table"

    # Add heading
    sheet['A1'] = "Rainfall Statistics"
    sheet['A1'].font = Font(size=20.5, bold=True)
    sheet.merge_cells('A1:B1')
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Add data
    data = [
        ("Month", "Rainfall (mm)"),
        ("February", 286),
        ("March", 281),
        ("April", 291),
        ("May", 158),
        ("June", 181),
        ("July", 133),
        ("August", 191)
    ]

    for row, (month, rainfall) in enumerate(data, start=3):
        sheet[f'A{row}'] = month
        sheet[f'B{row}'] = rainfall

    # Calculate statistics
    sheet['A10'] = "Average"
    sheet['B10'] = f"=AVERAGE(B4:B10)"
    sheet['A11'] = "Maximum"
    sheet['B11'] = f"=MAX(B4:B10)"
    sheet['A12'] = "Minimum"
    sheet['B12'] = f"=MIN(B4:B10)"

    # Apply borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet['A3:B12']:
        for cell in row:
            cell.border = thin_border

    # Create charts in Table sheet
    create_chart(sheet, "bar", "Rainfall Statistics over 7 Months", "A14")
    create_chart(sheet, "pie", "Rainfall Distribution", "A30")

    # Create a new sheet for graphs
    graph_sheet = wb.create_sheet("Graphs")

    # Create charts in Graphs sheet
    create_chart(graph_sheet, "bar", "Rainfall Statistics over 7 Months", "A1")
    create_chart(graph_sheet, "pie", "Rainfall Distribution", "A17")

    # Save the workbook
    wb.save("rainfall_statistics.xlsx")

if __name__ == "__main__":
    create_rainfall_excel()